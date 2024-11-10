from exam_class import ExamTargetClass
import openpyxl
import concurrent.futures
import warnings
from bs4 import XMLParsedAsHTMLWarning
import requests
import csv

# XMLParsedAsHTMLWarningを無視する
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

def save_intermediate_results(table, csv_path='intermediate_results.csv'):
    """
    中間結果をCSVに保存します。元の行番号と結果を合わせて保存し、再開時の整合性を保ちます。
    """
    with open(csv_path, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        
        # ヘッダーの書き込み（元のテーブルの列と追加する列）
        writer.writerow(table[0] + ["審査結果", "正解URL", "不足条文"])  # ヘッダー行を保持
        
        # 各行を保存
        for row in table[1:]:  # ヘッダーを除いたデータ部分
            processed_row = [
                str(item) if item is not None else "" for item in row
            ]
            writer.writerow(processed_row)


def system_validate(xlsx_path, new_xlsx_path):
    '''
    # エクセルをテーブルデータとして読み込む
    # 遵守事項掲載URLの列を取得
    # そのURLでone_testを実行
    # 審査結果をテーブルデータに追加（審査結果、正解URL、不足条文の3列を追加）
    # 追加部分を複製したエクセルに書き込む
    '''
    # メイン処理
    table = read_xlsx(xlsx_path)
    url_column = get_url_column(table)
    table = add_result_to_table(table, url_column)
    write_xlsx(table, new_xlsx_path)

# エクセルをテーブルデータとして読み込む
def read_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    table = []
    for row in ws.iter_rows(values_only=True):
        table.append(row)
    return table

# 遵守事項掲載URLの列を取得
def get_url_column(table):
    url_column = []
    for row in table:
        url_column.append(row[3])
    return url_column

# そのURLでone_testを実行
def test_url(url):
    base_json_path = 'base.json'
    exam = ExamTargetClass(url, base_json_path)
    result = exam.exam_all_urls()
    final_status = result["final_status"]
    print("url: ", url)
    if final_status == 1:
        result["final_status"] = 'OK'
        print("審査結果：OK")
    elif final_status == 2:
        result["final_status"] = '内容不備あり'
        print("審査結果：内容不備あり")
    else:
        result["final_status"] = '閲覧不可・動線不明'
        print("審査結果：閲覧不可・動線不明")
    print("=====================================")
    return result

# 審査結果をテーブルデータに追加（審査結果、正解URL、不足条文の3列を追加）
"""
def add_result_to_table(table, url_column):
    #for i in range(1, len(table)):
    for i in range(1, 10):
        print("URL: ", url_column[i])
        # urlが空欄の場合はスキップ
        if not url_column[i]:
            print("URLが空欄です")
            print("=====================================")
            continue
        result = test_url(url_column[i])
        final_status = result["final_status"]
        links = result["links"]
        missing_clauses = result["missing_clauses"]

        table[i] += (final_status, links, missing_clauses)
        print(final_status, links, missing_clauses)
    return table
"""
def process_url(url):
    if not url:
        print("URLが空欄です")
        print("=====================================")
        return {"final_status": None, "links": None, "missing_clauses": None}
    
    try:
        # タイムアウトを5秒に設定
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # ステータスコードが200以外なら例外を発生

        # 審査を行う
        result = test_url(url)
    except requests.exceptions.Timeout:
        print(f"URL {url} の処理がタイムアウトしました")
        return {"final_status": "Timeout", "links": None, "missing_clauses": None}
    except requests.exceptions.RequestException as e:
        print(f"URL {url} のリクエスト中にエラーが発生しました: {e}")
        return {"final_status": "Error", "links": None, "missing_clauses": None}

    return result

def add_result_to_table(table, url_column):
    # 並列処理を使用してURLごとに審査を実行
    with concurrent.futures.ProcessPoolExecutor(max_workers=8) as executor:
        future_to_url = {executor.submit(process_url, url): url for url in url_column[1:670]}
        
        results = []
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                result = future.result(timeout=10)
                results.append(result)
            except concurrent.futures.TimeoutError:
                print(f"URL {url} の処理がタイムアウトしました")
                results.append({"final_status": "Timeout", "links": None, "missing_clauses": None})
            except Exception as e:
                print(f"URL {url} の処理中にエラーが発生しました: {e}")
                # 結果が取得できなかった場合にNoneなどで代用
                results.append({"final_status": "Error", "links": [], "missing_clauses": []})

                table[i] += (result["final_status"], result["links"], result["missing_clauses"])
                save_intermediate_results(table)

    # 結果をテーブルに追加
    for i, result in enumerate(results, start=1):
        final_status = result["final_status"]
        links = result["links"]
        missing_clauses = result["missing_clauses"]
        table[i] += (final_status, links, missing_clauses)
        print("審査結果:", final_status, "リンク:", links, "不足条文:", missing_clauses)

    return table

# 追加部分を複製したエクセルに書き込む
def write_xlsx(table, xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active


    def flatten_list(nested_list):
        # 入れ子リストをフラット化
        flat_list = []
        for item in nested_list:
            if isinstance(item, list):
                flat_list.extend(flatten_list(item))
            else:
                flat_list.append(str(item))  # 文字列に変換
        return flat_list

    for row in table:
        # ワークシートの8-10列目を明示的に指定して審査結果、正解URL、不足条文を追加
        processed_row = [
            ",".join(flatten_list(item)) if isinstance(item, list) else item for item in row
        ]
        ws.append(processed_row)
    wb.save(xlsx_path)



def compare_result(manual_xlsx, system_xlsx):
    '''
    手動審査とシステム審査の2つの審査結果を比較する
    手動審査を真値として、TP, FP, TN、FNをカウントする
    注意すべきなのは、「追記・削除あり、記載ミスあり、初版」の３つは、システム審査で内容不備ありとして扱う
    '''
    # 手動審査の結果を読み込む
    manual_table = read_xlsx(manual_xlsx)
    manual_url_column = get_url_column(manual_table)
    manual_result_column = []
    for row in manual_table:
        manual_result_column.append(row[7])
    # システム審査の結果を読み込む
    system_table = read_xlsx(system_xlsx)
    system_url_column = get_url_column(system_table)
    system_result_column = []
    for row in system_table:
        system_result_column.append(row[7])

    # TP, FP, TN, FNをカウントする
    TP = 0
    FP = 0
    TN = 0
    FN = 0
    for i in range(1, len(manual_result_column)):
        if manual_result_column[i] == '遵守' and system_result_column[i] == '遵守':
            TP += 1
        elif manual_result_column[i] == '遵守' and system_result_column[i] != '遵守':
            FN += 1
        elif manual_result_column[i] != '遵守' and system_result_column[i] == '遵守':
            FP += 1
        else:
            TN += 1



if __name__ == '__main__':
    xlsx_path = '遵守宣言一覧.xlsx'
    new_xlsx_path = 'after.xlsx'
    system_validate(xlsx_path, new_xlsx_path)